from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, Side

from src.misc import *
from src.slack import *

class statsModes(Enum):
    # Minimum - Remove empty entries for user and channel, do not add entries for blank days
    MIN = 1
    # Medium - Remove empty entries for user and channel, create entries for blank days
    MEDIUM = 2
    # Full - Remove no entries, create entries for blank days
    FULL = 3

    def __le__(self, other):
        return self.value <= other.value
    def __ge__(self, other):
        return self.value >= other.value

class stats():
    SUBTYPES_WHITELIST = ('reminder_add',
                          'me_message',
                          'file_comment',
                          'file_mention',
                          'file_share')

    # Default stats mode
    mode = statsModes.MEDIUM

    def __init__(self, slack: slackData):
        self.slack = slack

        # Initialise blank maps and then calculate stats
        self.user_count = {}
        for u in self.slack.metadata.users:
            self.user_count[u] = 0
        self.channel_count = {}
        self.day_count = {}

        self.__calculateStats()
        self.tot_messages = sum(self.user_count.values())

    def exportPostStats(self):
        wb = Workbook()
        out_file = io.stats_dir + "Post stats.xlsx"

        # Create two sheets for users and channels
        wb_users = wb.active
        wb_users.title = "Users"
        wb_channels = wb.create_sheet(title="Channels")
        wb_days = wb.create_sheet(title="Days")
        self.__createColumns(wb_channels)
        self.__createColumns(wb_days)
        self.__createColumns(wb_users)

        # Add data
        sorted_dates = sorted(self.day_count.keys())

        self.__addStats(wb_channels, self.slack.metadata.channels, self.channel_count, prefix='#')
        self.__addStats(wb_days, sorted_dates, self.day_count, entry_num_format=misc.dateMode.toExcel())
        self.__addStats(wb_users, self.slack.metadata.users, self.user_count)

        # Add number of messages and export date range
        self.__addStatsInfo(wb_channels)
        self.__addStatsInfo(wb_days)
        self.__addStatsInfo(wb_users)

        # Make it pretty
        self.__adjustColumnWidth(wb_channels)
        self.__adjustColumnWidth(wb_days)
        self.__adjustColumnWidth(wb_users)

        # Save
        log.log(logModes.LOW, "Exporting statistics to '" + out_file + "'")
        io.ensureDir(io.stats_dir)

        try:
            wb.save(filename=out_file)
        except PermissionError:
            log.log(logModes.ERROR, "Could not write to file: Permission Denied")

    def __calculateStats(self):
        log.log(logModes.LOW, "Calculating statistics")

        # Iterate over history, update maps
        for channel in self.slack.metadata.channels:
            channel_count = 0

            for msg in self.slack.channel_data[channel]:
                # Get timestamp
                d = datetime.datetime.fromtimestamp(float(msg['ts'])).date()

                # Handle special cases
                if 'subtype' in msg:
                    if msg['subtype'] not in stats.SUBTYPES_WHITELIST:
                        continue
                    elif msg['subtype'] == 'file_comment':  # Handle file_comment, since user data is in 'comment' field
                        msg = msg['comment']
                if 'user' not in msg:
                    continue

                if self.slack.metadata.isDefinitelyUser(msg):
                    channel_count += 1
                    self.user_count[self.slack.metadata.getUserName(msg)] += 1

                    if d in self.day_count:
                        self.day_count[d] += 1
                    else:
                        self.day_count[d] = 1

            self.channel_count[channel] = channel_count

        # Add blank days
        if stats.mode >= statsModes.MEDIUM:
            min_date = min(x for x in self.day_count)
            max_date = max(x for x in self.day_count)

            for date in misc.daterange(min_date, max_date):
                if date not in self.day_count:
                    self.day_count[date] = 0

        # Delete empty keys from users and channels https://stackoverflow.com/a/15158637
        if stats.mode <= statsModes.MEDIUM:
            self.channel_count = {k: v for k, v in self.channel_count.items() if v > 0}
            self.user_count = {k: v for k, v in self.user_count.items() if v > 0}

    def __addStatsInfo(self, ws):
        # Tot messages
        ws['E1'] = "Total messages:"
        ws['E1'].alignment = Alignment(horizontal='right')
        ws['F1'] = int(self.tot_messages)
        ws['F1'].alignment = Alignment(horizontal='center')

    def __addStats(self, ws, values_ls: list, values_map: dict, prefix:str='', entry_num_format=None):
        i = 0
        for val in values_ls:
            if val not in values_map:
                continue

            # Calculate info
            messages = values_map[val]
            percentage = messages / self.tot_messages
            percentage = round(percentage, 3)

            # Save into workbook
            if not prefix == '':
                ws.cell(row=(i + 2), column=1).value = prefix + val
            else:
                ws.cell(row=(i + 2), column=1).value = val
            if entry_num_format is not None:
                ws.cell(row=(i + 2), column=1).number_format = entry_num_format

            ws.cell(row=(i + 2), column=2).value = messages
            ws.cell(row=(i + 2), column=2).number_format = "0"
            ws.cell(row=(i + 2), column=3).value = percentage
            ws.cell(row=(i + 2), column=3).number_format = "0.0%"

            # Left align 1st column, Center 2nd and 3rd column
            ws.cell(row=(i + 2), column=1).alignment = Alignment(horizontal='left')
            ws.cell(row=(i + 2), column=2).alignment = Alignment(horizontal='center')
            ws.cell(row=(i + 2), column=3).alignment = Alignment(horizontal='center')

            i += 1

    def __createColumns(self, ws):
        # Add column names
        ws['A1'] = ws.title[:-1]
        ws['B1'] = "Messages"
        ws['C1'] = "Percentage"

        # Align horizontally messages and percentage aligned
        ws['B1'].alignment = Alignment(horizontal='center')
        ws['C1'].alignment = Alignment(horizontal='center')

        # Make column titles bold and underlined
        ws['A1'].font = Font(bold=True)
        ws['B1'].font = Font(bold=True)
        ws['C1'].font = Font(bold=True)
        ws['A1'].border = Border(bottom=Side(border_style='thin'))
        ws['B1'].border = Border(bottom=Side(border_style='thin'))
        ws['C1'].border = Border(bottom=Side(border_style='thin'))

        # Add filter rules
        ws.auto_filter.ref = "A1:C" + str(len(self.slack.metadata.users))

    def __adjustColumnWidth(self, ws):
        # modified from https://stackoverflow.com/a/39530676

        for col in ws.columns:
            max_length = 0
            column = col[0].column  # Get the column name
            for cell in col:
                try:  # Necessary to avoid error on empty cells
                    if type(cell.value) is datetime.datetime:
                        cell_str = cell.value.strftime(misc.dateMode.value)
                    else:
                        cell_str = str(cell.value)

                    if len(cell_str) > max_length:
                        max_length = len(cell_str)
                except:
                    pass

            adjusted_width = (max_length + 0.5) * 1.2
            ws.column_dimensions[column].width = adjusted_width